from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models import Customer, ExportRun, Order, Product
from app.schemas.analytics import AnalyticsFilters
from app.services.exports import create_export


def make_session() -> Session:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    return Session(engine)


def seed(db: Session):
    db.add(
        Customer(
            mercos_id="c1",
            name="Cliente",
            active=True,
        )
    )
    db.add(
        Order(
            mercos_id="o1",
            number="1",
            customer_mercos_id="c1",
            status="2",
            issued_at=datetime(2026, 8, 12, 12, tzinfo=timezone.utc),
            total=Decimal("123.45"),
        )
    )
    db.add(
        Product(
            mercos_id="p1",
            name="Produto",
            list_price=Decimal("10"),
            stock=Decimal("2"),
            active=True,
        )
    )
    db.commit()


def test_filtered_csv_and_xlsx_exports_record_runs():
    with make_session() as db:
        seed(db)
        filters = AnalyticsFilters(period="all", customerIds=["c1"])

        csv_path, _, _, csv_run = create_export(
            db,
            username="admin",
            report="orders",
            export_format="csv",
            filters=filters,
        )
        xlsx_path, _, _, xlsx_run = create_export(
            db,
            username="admin",
            report="orders",
            export_format="xlsx",
            filters=filters,
        )
        inventory_path, _, _, inventory_run = create_export(
            db,
            username="admin",
            report="inventory",
            export_format="csv",
            filters=AnalyticsFilters(period="all"),
        )
        try:
            csv_text = csv_path.read_text(encoding="utf-8-sig")
            assert "customerName" in csv_text
            assert "Cliente" in csv_text
            workbook = load_workbook(xlsx_path, read_only=True)
            rows = list(workbook["orders"].iter_rows(values_only=True))
            assert rows[0][0] == "id"
            assert rows[1][0] == "o1"
            workbook.close()
            runs = list(db.scalars(select(ExportRun).order_by(ExportRun.id)))
            assert [run.id for run in runs] == [
                csv_run,
                xlsx_run,
                inventory_run,
            ]
            assert all(run.status == "success" and run.rows == 1 for run in runs)
            assert runs[0].filters["customerIds"] == ["c1"]
        finally:
            csv_path.unlink(missing_ok=True)
            xlsx_path.unlink(missing_ok=True)
            inventory_path.unlink(missing_ok=True)
