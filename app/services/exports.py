from __future__ import annotations

import csv
from datetime import date, datetime, timezone
from decimal import Decimal
import json
from pathlib import Path
import tempfile
from typing import Iterator, Literal

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models import ExportRun
from app.schemas.analytics import AnalyticsFilters
from app.services.analytics_filters import applied_filters
from app.services.analytics_v2 import (
    customers_page,
    inventory_page,
    orders_page,
    products_page,
    sellers_page,
)


Report = Literal["orders", "products", "customers", "sellers", "inventory"]
Format = Literal["csv", "xlsx"]


REPORT_CONFIG = {
    "orders": (orders_page, "issued_at", "desc"),
    "products": (products_page, "revenue", "desc"),
    "customers": (customers_page, "revenue", "desc"),
    "sellers": (sellers_page, "revenue", "desc"),
    "inventory": (inventory_page, "stock_value", "desc"),
}


def _safe_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=_safe_value)
    if value is None:
        return ""
    return value


def _iter_rows(
    db: Session,
    report: Report,
    filters: AnalyticsFilters,
) -> Iterator[dict]:
    function, sort, order = REPORT_CONFIG[report]
    page = 1
    while True:
        result = function(
            db,
            filters,
            page=page,
            page_size=100,
            search=None,
            sort=sort,
            order=order,
        )
        yield from result["items"]
        if page >= result["totalPages"]:
            return
        page += 1


def _create_run(
    db: Session,
    *,
    username: str,
    report: Report,
    export_format: Format,
    filters: AnalyticsFilters,
) -> ExportRun:
    run = ExportRun(
        username=username,
        report=report,
        format=export_format,
        status="running",
        filters=applied_filters(filters),
        started_at=datetime.now(timezone.utc),
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    return run


def _finish_run(
    db: Session,
    run_id: int,
    *,
    status: str,
    rows: int,
    error: str | None = None,
) -> None:
    run = db.get(ExportRun, run_id)
    if run is None:
        return
    run.status = status
    run.finished_at = datetime.now(timezone.utc)
    run.rows = rows
    run.error = error
    db.add(run)
    db.commit()


def create_export(
    db: Session,
    *,
    username: str,
    report: Report,
    export_format: Format,
    filters: AnalyticsFilters,
) -> tuple[Path, str, str, int]:
    run = _create_run(
        db,
        username=username,
        report=report,
        export_format=export_format,
        filters=filters,
    )
    suffix = f".{export_format}"
    file = tempfile.NamedTemporaryFile(
        prefix=f"xnamai-{report}-",
        suffix=suffix,
        delete=False,
    )
    path = Path(file.name)
    file.close()
    row_count = 0
    try:
        rows = _iter_rows(db, report, filters)
        first = next(rows, None)
        headers = list(first) if first else []
        if export_format == "csv":
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=headers)
                if headers:
                    writer.writeheader()
                    writer.writerow(
                        {key: _safe_value(first[key]) for key in headers}
                    )
                    row_count = 1
                    for row in rows:
                        writer.writerow(
                            {key: _safe_value(row.get(key)) for key in headers}
                        )
                        row_count += 1
            content_type = "text/csv; charset=utf-8"
        else:
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet(title=report[:31])
            if headers:
                sheet.append(headers)
                sheet.append([_safe_value(first[key]) for key in headers])
                row_count = 1
                for row in rows:
                    sheet.append([_safe_value(row.get(key)) for key in headers])
                    row_count += 1
            workbook.save(path)
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        _finish_run(db, run.id, status="success", rows=row_count)
        download_name = (
            f"xnamai-{report}-{datetime.now().date().isoformat()}.{export_format}"
        )
        return path, content_type, download_name, run.id
    except Exception as exc:
        path.unlink(missing_ok=True)
        _finish_run(
            db,
            run.id,
            status="error",
            rows=row_count,
            error=str(exc)[:1000],
        )
        raise
